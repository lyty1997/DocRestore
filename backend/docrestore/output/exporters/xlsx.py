# Copyright 2026 @lyty1997
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""xlsx 导出器（D4）：``document.md`` → Excel，走 openpyxl。

`document.md` 里的表格**一律是 HTML ``<table>``**（LLM prompt 默认保留 HTML 原样，
`table_dedup.py` 也按 ``<tr>/<td>`` 去重）。HTML 表本身携带行列与 ``colspan/rowspan``——
它就是结构化 IR，无需 pipeline 旁路。本导出器在**下载环节**解析每个 ``<table>`` →
单元格矩阵 + 合并区，openpyxl **每表一 sheet**；纯数字单元格转数值；合并区
``merge_cells``。文档无表 → 退化为单 ``Document`` sheet（每非空 markdown 行一行，
产物非空便于派生断言）。

openpyxl 是纯 Python 依赖（无系统库）：**惰性导入** fail-closed
（:class:`ExportToolUnavailable`）——注册表启动时导入本模块，顶层不 import openpyxl。
详见 ``docs/zh/export-mode.md`` §9.1。
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import TypeAlias

from docrestore.output.exporters.base import (
    ExportFailed,
    ExportToolUnavailable,
)
from docrestore.output.exporters.html_table import RawCell
from docrestore.output.exporters.html_table import build_grid as _build_grid
from docrestore.output.exporters.html_table import parse_tables as _parse_tables

logger = logging.getLogger(__name__)

#: 纯整数（禁前导 0，避免电话/编号被转成数字丢前导 0）
_INT_RE = re.compile(r"-?[1-9]\d*|0")
#: 纯小数
_FLOAT_RE = re.compile(r"-?\d+\.\d+")
#: 退化文本 sheet 的最大行数（挡超大文档失控写盘）
_MAX_TEXT_ROWS = 5000

#: 单元格值的联合类型（openpyxl 接受 str/int/float）
CellValue: TypeAlias = str | int | float


def _coerce_cell(text: str) -> CellValue:
    """纯数字单元格 → 数值（``"100"``→100），否则保留字符串。"""
    if _INT_RE.fullmatch(text):
        return int(text)
    if _FLOAT_RE.fullmatch(text):
        return float(text)
    return text


class XlsxExporter:
    """``document.md`` → xlsx（openpyxl，每表一 sheet）。"""

    suffix = "xlsx"
    tool = "openpyxl"

    def ensure_available(self) -> None:
        """openpyxl 不可导入 → fail-closed。"""
        try:
            import openpyxl  # noqa: F401, PLC0415 — 惰性导入仅做可用性探测
        except ImportError as exc:
            raise ExportToolUnavailable(self.tool) from exc

    def export(
        self,
        doc_md: Path,
        assets_dir: Path,  # noqa: ARG002 — xlsx 不嵌图，仅消费表格文本
        out_path: Path,
    ) -> None:
        """解析 ``doc_md`` 的 HTML 表 → openpyxl 工作簿；无表则落正文 sheet。"""
        try:
            import openpyxl  # noqa: PLC0415 — 惰性导入，避免缺依赖时整体起不来
        except ImportError as exc:
            raise ExportToolUnavailable(self.tool) from exc

        markdown = doc_md.read_text(encoding="utf-8")
        tables = _parse_tables(markdown)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            workbook = openpyxl.Workbook()
            default_sheet = workbook.active  # 去掉默认空 sheet
            if default_sheet is not None:
                workbook.remove(default_sheet)
            if tables:
                for idx, rows in enumerate(tables, 1):
                    sheet = workbook.create_sheet(f"Table {idx}")
                    _fill_table_sheet(sheet, rows)
            else:
                _fill_text_sheet(workbook.create_sheet("Document"), markdown)
            workbook.save(str(out_path))
        except Exception as exc:  # noqa: BLE001 — openpyxl 异常统一 fail-closed
            raise ExportFailed(self.tool, self.suffix, str(exc)[:500]) from exc


def _fill_table_sheet(sheet: object, rows: list[list[RawCell]]) -> None:
    """把一张表写入 openpyxl sheet（含数值转换与合并区）。"""
    cells, merges = _build_grid(rows)
    for (r, c), text in cells.items():
        sheet.cell(row=r + 1, column=c + 1, value=_coerce_cell(text))  # type: ignore[attr-defined]
    for r1, c1, r2, c2 in merges:
        sheet.merge_cells(  # type: ignore[attr-defined]
            start_row=r1 + 1, start_column=c1 + 1,
            end_row=r2 + 1, end_column=c2 + 1,
        )


def _fill_text_sheet(sheet: object, markdown: str) -> None:
    """无表退化：每非空 markdown 行写入 A 列一行（产物非空）。"""
    row = 1
    truncated = False
    for line in markdown.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        sheet.cell(row=row, column=1, value=stripped)  # type: ignore[attr-defined]
        row += 1
        if row > _MAX_TEXT_ROWS:
            truncated = True
            break
    if truncated:
        # 不再静默截断：记 warning + 在末尾追加可见标记，避免"看着完整其实缺尾"。
        logger.warning(
            "xlsx 无表退化 sheet 正文超 %d 行已截断，余下行未导出", _MAX_TEXT_ROWS,
        )
        sheet.cell(  # type: ignore[attr-defined]
            row=row, column=1,
            value=f"（已截断：正文超过 {_MAX_TEXT_ROWS} 行，余下行未导出）",
        )
