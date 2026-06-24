# Copyright 2026 @lyty1997
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0

"""Epic D · D4 xlsx 导出（openpyxl，解析 HTML 表）测试。

- 纯解析 / 数值转换 / 网格合并：不依赖 openpyxl，恒可跑。
- 缺 openpyxl fail-closed：模拟 ImportError → ``ExportToolUnavailable``。
- 真导出（需 openpyxl）：含 ``colspan/rowspan`` 的 md → xlsx，openpyxl 读回
  **从输入派生**的单元格文字 / 合并区 / 数值类型；无表 → 单 ``Document`` sheet。
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pytest

from docrestore.output.exporters.base import ExportToolUnavailable
from docrestore.output.exporters.xlsx import (
    XlsxExporter,
    _build_grid,
    _coerce_cell,
    _parse_tables,
)


def _openpyxl_ready() -> bool:
    return importlib.util.find_spec("openpyxl") is not None


openpyxl_required = pytest.mark.skipif(
    not _openpyxl_ready(), reason="openpyxl 不可用",
)

#: 从输入派生的关键内容（构造输入，断言由此派生）
_HEADING = "季度营收汇总报告"
_CELL = "华东区域条目甲"
_NUM = 1280
_TABLE_MD = f"""# {_HEADING}

<table border=1>
<tr><th>区域</th><th colspan="2">营收</th></tr>
<tr><td rowspan="2">{_CELL}</td><td>Q1</td><td>{_NUM}</td></tr>
<tr><td>Q2</td><td>1530</td></tr>
</table>

正文段落，与表格无关。
"""


def _build_doc(doc_dir: Path, markdown: str) -> Path:
    (doc_dir / "document.md").write_text(markdown, encoding="utf-8")
    (doc_dir / "images").mkdir()
    return doc_dir / "document.md"


class TestParseAndCoerce:
    """纯函数：表抽取 / 网格合并 / 数值转换（无依赖）。"""

    def test_parse_one_table_three_rows(self) -> None:
        tables = _parse_tables(_TABLE_MD)
        assert len(tables) == 1
        assert len(tables[0]) == 3  # 表头 + 2 数据行

    def test_grid_merges_from_spans(self) -> None:
        tables = _parse_tables(_TABLE_MD)
        cells, merges = _build_grid(tables[0])
        # 表头单元格文字派生自输入
        assert cells[(0, 0)] == "区域"
        assert cells[(1, 0)] == _CELL
        # colspan=2 → 合并 (0,1)-(0,2)；rowspan=2 → 合并 (1,0)-(2,0)
        assert (0, 1, 0, 2) in merges
        assert (1, 0, 2, 0) in merges

    def test_no_table_returns_empty(self) -> None:
        assert _parse_tables("# 纯标题\n\n正文无表。\n") == []

    @pytest.mark.parametrize(
        ("text", "expected"),
        [("1280", 1280), ("0", 0), ("-3", -3), ("1.5", 1.5),
         ("007", "007"), ("abc", "abc")],
    )
    def test_coerce_cell(self, text: str, expected: object) -> None:
        result = _coerce_cell(text)
        assert result == expected
        assert type(result) is type(expected)


class TestXlsxFailClosed:
    """缺 openpyxl → ensure_available fail-closed。"""

    def test_missing_openpyxl_raises(
        self, monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        monkeypatch.setitem(sys.modules, "openpyxl", None)
        with pytest.raises(ExportToolUnavailable):
            XlsxExporter().ensure_available()


@openpyxl_required
class TestXlsxExport:
    """真导出（需 openpyxl）：派生单元格 / 合并区 / 数值 + 无表退化。"""

    def test_roundtrip_table(self, tmp_path: Path) -> None:
        import openpyxl

        doc_md = _build_doc(tmp_path, _TABLE_MD)
        out = tmp_path / ".exports" / "out.xlsx"

        XlsxExporter().export(doc_md, tmp_path / "images", out)

        workbook = openpyxl.load_workbook(out)
        assert "Table 1" in workbook.sheetnames
        sheet = workbook["Table 1"]
        values = [
            sheet.cell(r, c).value
            for r in range(1, sheet.max_row + 1)
            for c in range(1, sheet.max_column + 1)
        ]
        assert _CELL in values  # 派生：跨行单元格文字
        assert _NUM in values  # 派生：数值
        assert isinstance(sheet.cell(2, 3).value, int)  # 数值单元格为整数
        merges = {str(m) for m in sheet.merged_cells.ranges}
        assert "B1:C1" in merges  # colspan=2
        assert "A2:A3" in merges  # rowspan=2

    def test_no_table_degenerate_text_sheet(self, tmp_path: Path) -> None:
        import openpyxl

        heading = "无表文档标题乙"
        doc_md = _build_doc(tmp_path, f"# {heading}\n\n正文一段。\n")
        out = tmp_path / ".exports" / "out.xlsx"

        XlsxExporter().export(doc_md, tmp_path / "images", out)

        workbook = openpyxl.load_workbook(out)
        assert workbook.sheetnames == ["Document"]
        col_a = [
            workbook["Document"].cell(r, 1).value
            for r in range(1, workbook["Document"].max_row + 1)
        ]
        assert f"# {heading}" in col_a  # 派生：正文行落 A 列
